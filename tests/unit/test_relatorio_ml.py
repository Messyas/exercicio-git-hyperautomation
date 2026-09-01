from pathlib import Path
import openpyxl
import pytest

from src.relatorio_executivo import gerar_excel_consolidado
from src.servico_validacao import RegistroValidado
from src.item_processor import MLDecision
from src.operational_indicators import calcular_indicadores


def test_excel_9_abas_e_aba_decisoes_ml(tmp_path):
    """Valida se o Excel consolidado é gerado com exatamente 9 abas e se a aba 'Decisões de ML' está correta."""
    validados = [
        RegistroValidado("L-1", "Prod A", "L1", "A", "APROVADO", "Tech 1", "01/01/2026", "", "01/01/2026", classificacao="Válido"),
        RegistroValidado("L-2", "Prod B", "L2", "B", "PENDENTE", "Tech 2", "01/01/2026", "", "01/01/2026", classificacao="Ambíguo"),
        RegistroValidado("L-3", "Prod C", "L3", "C", "EM AJUSTE", "Tech 3", "01/01/2026", "Ajustando", "01/01/2026", classificacao="Ambíguo"),
    ]
    indicadores = calcular_indicadores(validados)

    decisoes_ml = [
        MLDecision("2026-01-01T10:00:00Z", "L-2", "PENDENTE", "B", False, "revisar", 0.75, "media", "REVISAR", 12.5, True, False, "rf-lotes-1.0.0"),
        MLDecision("2026-01-01T10:00:01Z", "L-3", "EM AJUSTE", "C", True, "valido_automatico", 0.89, "alta", "VALIDO_AUTOMATICO", 15.2, True, False, "rf-lotes-1.0.0"),
    ]

    excel_path = tmp_path / "test_9_abas.xlsx"
    gerar_excel_consolidado(validados, indicadores, excel_path, decisoes_ml=decisoes_ml)

    assert excel_path.exists()
    wb = openpyxl.load_workbook(excel_path)

    nombres_esperados = [
        "Resumo",
        "Todos",
        "Válidos",
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Ranking de Regras",
        "Dicionário",
        "Decisões de ML",
    ]
    assert wb.sheetnames == nombres_esperados

    # Valida 9ª aba
    ws_ml = wb["Decisões de ML"]
    headers = [cell.value for cell in ws_ml[1]]
    assert "Timestamp" in headers
    assert "Lote" in headers
    assert "Classe predita" in headers
    assert "Probabilidade" in headers
    assert "Ação final" in headers

    # Quantidade de linhas de dados == quantidade de decisões ML (2)
    assert ws_ml.max_row - 1 == len(decisoes_ml)

    # Formato da probabilidade
    prob_cell = ws_ml.cell(row=2, column=headers.index("Probabilidade") + 1)
    assert prob_cell.value == 0.75
    assert prob_cell.number_format == "0.00%"
