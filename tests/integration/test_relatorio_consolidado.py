"""Teste de integração consolidado do Dashboard Executivo (Aula 24)."""

from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pytest
from openpyxl.chart import DoughnutChart, LineChart

from src.relatorio_executivo import executar_pipeline_dashboard
import src.operational_indicators as op_ind_module


@pytest.mark.integration
def test_relatorio_consolidado_8_abas_e_markdown(tmp_path: Path):
    caminho_entrada = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "samples"
        / "inspecao_lotes_10dias_sem gabarito.xlsx"
    )

    # Monitorar chamadas para calcular_indicadores
    chamadas_calcular = 0
    orig_calcular = op_ind_module.calcular_indicadores

    def spy_calcular(*args, **kwargs):
        nonlocal chamadas_calcular
        chamadas_calcular += 1
        return orig_calcular(*args, **kwargs)

    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr("src.relatorio_executivo.calcular_indicadores", spy_calcular)

    try:
        resultados = executar_pipeline_dashboard(caminho_entrada, tmp_path)
    finally:
        monkeypatch.undo()

    # Comprovar chamada única a calcular_indicadores
    assert chamadas_calcular == 1

    caminho_excel = resultados["excel"]
    caminho_md = resultados["markdown"]
    caminho_pdf = resultados["pdf"]
    caminho_log = resultados["log"]

    assert caminho_excel.exists()
    assert caminho_md.exists()
    assert caminho_pdf.exists()
    assert caminho_log.exists()

    # Inspeção do Workbook Excel
    wb = openpyxl.load_workbook(caminho_excel)
    abas_esperadas = [
        "Resumo",
        "Todos",
        "Válidos",
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Ranking de Regras",
        "Dicionário",
        "Decisões de ML",
    ]
    assert wb.sheetnames == abas_esperadas


    # Aba Todos
    ws_todos = wb["Todos"]
    registros_todos = list(ws_todos.iter_rows(values_only=True))
    assert len(registros_todos) == 251  # Cabeçalho + 250 linhas

    # Abas de detalhe: sem mistura de classificações
    ws_validos = wb["Válidos"]
    val_rows = list(ws_validos.iter_rows(values_only=True))
    assert len(val_rows) == 151
    for r in val_rows[1:]:
        assert r[10] == "Válido"  # Coluna Classificação

    ws_div = wb["Divergências"]
    div_rows = list(ws_div.iter_rows(values_only=True))
    assert len(div_rows) == 51
    for r in div_rows[1:]:
        assert r[10] == "Divergência"

    ws_amb = wb["Ambíguos"]
    amb_rows = list(ws_amb.iter_rows(values_only=True))
    assert len(amb_rows) == 21
    for r in amb_rows[1:]:
        assert r[10] == "Ambíguo"

    ws_err = wb["Erros de Entrada"]
    err_rows = list(ws_err.iter_rows(values_only=True))
    assert len(err_rows) == 31
    for r in err_rows[1:]:
        assert r[10] == "Erro de Entrada"

    # Aba Resumo: 10 indicadores, valores constantes e referências de negócio.
    ws_resumo = wb["Resumo"]
    labels_resumo = [ws_resumo.cell(row=r, column=1).value for r in range(6, 16)]
    assert "Total de registros" in labels_resumo
    assert "Registros válidos" in labels_resumo
    assert "Divergências" in labels_resumo
    assert "Ambíguos" in labels_resumo
    assert "Erros de Entrada" in labels_resumo
    assert "Regra mais acionada" in labels_resumo
    assert "Taxa de qualidade da entrada" in labels_resumo
    assert "Taxa de revisão humana" in labels_resumo
    assert "Taxa de retrabalho" in labels_resumo
    assert "Ganho estimado de tempo" in labels_resumo

    valores_resumo = {
        ws_resumo.cell(row=r, column=1).value: (
            ws_resumo.cell(row=r, column=2).value,
            ws_resumo.cell(row=r, column=3).value,
            ws_resumo.cell(row=r, column=4).value,
        )
        for r in range(6, 16)
    }
    assert valores_resumo["Total de registros"] == (250, "-", "-")
    assert valores_resumo["Registros válidos"] == (150, 0.6, "Informativo")
    assert valores_resumo["Divergências"] == (50, 0.2, "Informativo")
    assert valores_resumo["Ambíguos"] == (20, 0.08, "Informativo")
    assert valores_resumo["Erros de Entrada"] == (30, 0.12, "Informativo")
    assert valores_resumo["Regra mais acionada"][0].startswith("RN06")
    assert valores_resumo["Regra mais acionada"][1] == 0.1
    assert valores_resumo["Taxa de qualidade da entrada"] == (0.88, 0.88, "> 80%")
    assert valores_resumo["Taxa de revisão humana"] == (0.08, 0.08, "< 15%")
    assert valores_resumo["Taxa de retrabalho"] == (0.2, 0.2, "< 6%")
    assert valores_resumo["Ganho estimado de tempo"] == ("437.5 min", "-", "-")

    # Verificar ausência de fórmulas nas células dos indicadores
    for r in range(6, 16):
        for col in range(1, 6):
            val = str(ws_resumo.cell(row=r, column=col).value or "")
            assert not val.startswith("="), f"Célula ({r},{col}) contém fórmula: {val}"

    # Gráficos nativos na aba Resumo: uma rosca com as quatro classificações
    # e uma linha com a evolução diária. A leitura do XML garante os intervalos.
    assert len(ws_resumo._charts) == 2
    assert any(isinstance(chart, DoughnutChart) for chart in ws_resumo._charts)
    assert any(isinstance(chart, LineChart) for chart in ws_resumo._charts)
    with ZipFile(caminho_excel) as arquivo_xlsx:
        graficos = [
            arquivo_xlsx.read(nome).decode("utf-8")
            for nome in arquivo_xlsx.namelist()
            if nome.startswith("xl/charts/") and nome.endswith(".xml")
        ]
    xml_rosca = next(xml for xml in graficos if "doughnutChart" in xml)
    assert "'Resumo'!$A$7:$A$10" in xml_rosca
    assert "'Resumo'!$B$7:$B$10" in xml_rosca
    assert "'Resumo'!$B$6:$B$10" not in xml_rosca

    # Aba Ranking de Regras
    ws_ranking = wb["Ranking de Regras"]
    ranking_rows = list(ws_ranking.iter_rows(values_only=True))
    header_ranking = ranking_rows[0]
    assert header_ranking == ("Posição", "Regra", "Nome", "Ocorrências", "% do total")
    primeira_regra = ranking_rows[1]
    assert primeira_regra[1] == "RN06"
    assert primeira_regra[3] == 25

    # Aba Dicionário
    ws_dic = wb["Dicionário"]
    dic_rows = list(ws_dic.iter_rows(values_only=True))
    termos_dic = [r[0] for r in dic_rows[1:] if r[0]]
    for regra_code in [f"RN0{i}" if i < 10 else f"RN{i}" for i in range(1, 13)]:
        assert any(regra_code in t for t in termos_dic), f"{regra_code} ausente do Dicionário"
    assert "Válido" in termos_dic
    assert "Divergência" in termos_dic
    assert "Ambíguo" in termos_dic
    assert "Erro de Entrada" in termos_dic

    # Inspeção do Markdown
    texto_md = caminho_md.read_text(encoding="utf-8")
    assert "# Resumo Executivo — Conferência de Lotes" in texto_md
    assert "## Visão Geral" in texto_md
    assert "## Indicadores Principais" in texto_md
    assert "## Destaque" in texto_md
    assert "## Ganho Estimado de Tempo" in texto_md
    assert "## Observação" in texto_md

    # Os números publicados no Markdown devem ser os mesmos do Excel.
    assert "**Total de registros processados:** 250" in texto_md
    assert "**Registros válidos:** 150 (60,0%)" in texto_md
    assert "**Divergências operacionais:** 50 (20,0%)" in texto_md
    assert "**Registros ambíguos (revisão humana):** 20 (8,0%)" in texto_md
    assert "**Erros de entrada:** 30 (12,0%)" in texto_md
    assert "88,0%" in texto_md
    assert "8,0%" in texto_md
    assert "20,0%" in texto_md
    assert "RN06" in texto_md
    assert "25 ocorrências" in texto_md
    assert "437,5 minutos" in texto_md
    assert "7h17min30s" in texto_md
    assert "estimativa didática" in texto_md

    # Sem identificadores internos no Markdown
    for termo_proibido in ["_percentual", "OperationalIndicators", "validar_registro", "RegistroValidado"]:
        assert termo_proibido not in texto_md
