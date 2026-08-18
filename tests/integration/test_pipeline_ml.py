from pathlib import Path
import json
import openpyxl
import pytest

from dashboard.main import executar_pipeline_dashboard


@pytest.mark.integration
def test_pipeline_dashboard_integracao_ml(tmp_path: Path):
    """Executa o pipeline completo do dashboard e verifica se os registros ambíguos geraram decisões na 9ª aba."""
    caminho_entrada = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "samples"
        / "inspecao_lotes_10dias_sem gabarito.xlsx"
    )

    resultados = executar_pipeline_dashboard(caminho_entrada, tmp_path)
    caminho_excel = resultados["excel"]
    assert caminho_excel.exists()

    wb = openpyxl.load_workbook(caminho_excel)
    assert "Decisões de ML" in wb.sheetnames

    ws_amb = wb["Ambíguos"]
    num_ambiguos = ws_amb.max_row - 1  # 20 na amostra

    ws_ml = wb["Decisões de ML"]
    num_decisoes = ws_ml.max_row - 1

    # Regra de completude: quantidade de decisões de ML deve ser exatamente igual aos ambíguos
    assert num_decisoes == num_ambiguos

    eventos = [
        json.loads(linha)
        for linha in resultados["log"].read_text(encoding="utf-8").splitlines()
    ]
    decisoes_logadas = [evento for evento in eventos if evento.get("event") == "ML_DECISION"]
    assert len(decisoes_logadas) == num_ambiguos
    assert all("lote_id" in evento and "latencia_ms" in evento for evento in decisoes_logadas)
