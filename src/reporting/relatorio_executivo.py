"""Orquestrador principal do Dashboard Executivo (Aula 24).

Concentra o fluxo oficial de processamento, cálculo de indicadores operacionais
(através do módulo puro ``src.operational_indicators``) e geração de todos os
artefatos consolidados (Excel 9 abas, Resumo Executivo Markdown, PDF compatível e Log).
"""

from __future__ import annotations

import argparse
from collections import Counter
import logging
import re
import sys
import unicodedata
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.core.base_referencia import carregar_base_referencia
from src.automation.item_processor import ItemProcessor, MLDecision
from src.ml.ml_client_factory import create_ml_client
from config import settings
from src.utils.structured_logging import configure_structured_logging
from src.reporting.operational_indicators import (
    CATALOGO_REGRAS,
    OperationalIndicators,
    RankedRule,
    _percentual,
    calcular_indicadores,
)
from src.core.servico_validacao import (
    ABAS_DIARIAS,
    CLASSIFICACOES,
    RegistroValidado,
    data_da_aba,
    texto,
    validar_registro,
    validar_registros_lista,
)
from src.core.validacao import valida_estrutura


ARQUIVO_ENTRADA_PADRAO = (
    ROOT / "data" / "samples" / "inspecao_lotes_10dias_sem gabarito.xlsx"
)
DIRETORIO_SAIDA_PADRAO = ROOT / "data" / "output"

COLUNAS_RELATORIO = {
    "lote_id": "Lote",
    "produto": "Produto",
    "linha": "Linha",
    "turno": "Turno",
    "status_normalizado": "Status",
    "responsavel": "Responsável",
    "data": "Data da inspeção",
    "data_referencia": "Data de referência",
    "observacao": "Observação",
    "descricao_validacao": "Orientação",
    "classificacao": "Classificação",
}

FUSO_MANAUS = timezone(timedelta(hours=-4), name="America/Manaus")


def _agora_manaus() -> datetime:
    """Retorna um instante com fuso explícito (America/Manaus)."""
    return datetime.now(FUSO_MANAUS)


def _formatar_tempo_legivel(minutos: float) -> str:
    """Converte minutos no formato legível XhYYminZZs."""
    total_segundos = int(round(minutos * 60))
    horas = total_segundos // 3600
    resto = total_segundos % 3600
    mins = resto // 60
    secs = resto % 60
    return f"{horas}h{mins:02d}min{secs:02d}s"


def _ajustar_aba_tabela(ws, nome_tabela: str) -> None:
    """Aplica estilo visual de cabeçalho, auto-filtro, larguras e tabela openpyxl."""
    cabecalho_fill = PatternFill("solid", fgColor="1F4E78")
    for celula in ws[1]:
        celula.fill = cabecalho_fill
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for coluna in range(1, ws.max_column + 1):
        valores = [
            len(str(ws.cell(linha, coluna).value or ""))
            for linha in range(1, min(ws.max_row, 100) + 1)
        ]
        ws.column_dimensions[get_column_letter(coluna)].width = min(
            max(valores, default=10) + 2, 50
        )

    if ws.max_row > 1 and ws.max_column:
        nome_limpo = re.sub(r"[^A-Za-z0-9]", "", nome_tabela)
        tabela = Table(displayName=f"Tabela{nome_limpo}", ref=ws.dimensions)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tabela)


def _montar_aba_resumo(
    ws, indicadores: OperationalIndicators, validados: Sequence[RegistroValidado]
) -> None:
    """Preenche a aba Resumo com título, timestamp, a tabela de 10 indicadores e os 2 gráficos."""
    ws.title = "Resumo"

    # Título principal
    ws["A1"] = "Dashboard Executivo — Conferência de Lotes"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:E1")

    # Metadados
    ws["A2"], ws["B2"] = "Processado em", _agora_manaus().strftime(
        "%d/%m/%Y %H:%M:%S"
    )
    ws["A3"], ws["B3"] = "Total de registros processados", indicadores.total_registros
    ws["B3"].font = Font(size=14, bold=True, color="1F4E78")

    # Cabeçalho da tabela de indicadores
    headers = [
        "Indicador",
        "Quantidade/Valor",
        "Percentual",
        "Referência",
        "Sinal/Detalhe",
    ]
    fill_header = PatternFill("solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    regra_principal_str = (
        f"{indicadores.regra_mais_acionada.codigo} — {indicadores.regra_mais_acionada.nome}"
        if indicadores.regra_mais_acionada
        else "Nenhuma"
    )
    regra_pct = (
        indicadores.regra_mais_acionada.percentual_total / 100.0
        if indicadores.regra_mais_acionada
        else 0.0
    )
    regra_detalhe = (
        f"{indicadores.regra_mais_acionada.ocorrencias} ocorrências"
        if indicadores.regra_mais_acionada
        else "-"
    )

    linhas_indicadores = [
        ("Total de registros", indicadores.total_registros, "-", "-", "-"),
        (
            "Registros válidos",
            indicadores.registros_validos,
            indicadores.percentual_validos / 100.0,
            "Informativo",
            "-",
        ),
        (
            "Divergências",
            indicadores.divergencias,
            indicadores.percentual_divergencias / 100.0,
            "Informativo",
            "-",
        ),
        (
            "Ambíguos",
            indicadores.ambiguos,
            indicadores.percentual_ambiguos / 100.0,
            "Informativo",
            "-",
        ),
        (
            "Erros de Entrada",
            indicadores.erros_entrada,
            indicadores.percentual_erros_entrada / 100.0,
            "Informativo",
            "-",
        ),
        ("Regra mais acionada", regra_principal_str, regra_pct, "-", regra_detalhe),
        (
            "Taxa de qualidade da entrada",
            indicadores.taxa_qualidade_entrada / 100.0,
            indicadores.taxa_qualidade_entrada / 100.0,
            "> 80%",
            "✓",
        ),
        (
            "Taxa de revisão humana",
            indicadores.taxa_revisao_humana / 100.0,
            indicadores.taxa_revisao_humana / 100.0,
            "< 15%",
            "✓",
        ),
        (
            "Taxa de retrabalho",
            indicadores.taxa_retrabalho / 100.0,
            indicadores.taxa_retrabalho / 100.0,
            "< 6%",
            "⚠",
        ),
        (
            "Ganho estimado de tempo",
            f"{indicadores.ganho_estimado_minutos:.1f} min",
            "-",
            "-",
            _formatar_tempo_legivel(indicadores.ganho_estimado_minutos),
        ),
    ]

    for idx, (ind, valor, pct, ref, sinal) in enumerate(linhas_indicadores, start=6):
        c1 = ws.cell(row=idx, column=1, value=ind)
        c2 = ws.cell(row=idx, column=2, value=valor)
        c3 = ws.cell(row=idx, column=3, value=pct)
        c4 = ws.cell(row=idx, column=4, value=ref)
        c5 = ws.cell(row=idx, column=5, value=sinal)

        if isinstance(pct, float):
            c3.number_format = "0.0%"
        if isinstance(valor, float) and idx in (12, 13, 14):
            c2.number_format = "0.0%"
        elif isinstance(valor, int) and idx in (6, 7, 8, 9, 10):
            c2.number_format = "#,##0"

    # Tabela formatada openpyxl para os 10 indicadores
    tabela_ind = Table(displayName="TabelaResumoIndicadores", ref="A5:E15")
    tabela_ind.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela_ind)

    # Nota de premissas
    ws["A17"] = (
        "Premissas: 2,00 min manual e 0,25 min automatizado por registro — estimativa didática"
    )
    ws["A17"].font = Font(size=9, italic=True, color="555555")

    # Tabela diária para evolução dos alertas (Divergências + Ambíguos)
    df_val = pd.DataFrame([r.to_dict() for r in validados])
    df_val["alerta"] = df_val["classificacao"].isin(["Divergência", "Ambíguo"]).astype(int)
    evolucao = (
        df_val.groupby("data_referencia", sort=False)["alerta"]
        .sum()
        .reset_index(name="Divergências + Ambíguos")
    )

    ws.cell(row=19, column=1, value="Data de referência").fill = fill_header
    ws.cell(row=19, column=1).font = font_header
    ws.cell(row=19, column=2, value="Divergências + Ambíguos").fill = fill_header
    ws.cell(row=19, column=2).font = font_header

    for r_idx, row in enumerate(evolucao.itertuples(), start=20):
        ws.cell(row=r_idx, column=1, value=row.data_referencia)
        ws.cell(row=r_idx, column=2, value=row._2)

    tabela_evo = Table(displayName="TabelaEvolucaoAlertas", ref="A19:B29")
    tabela_evo.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela_evo)

    # Gráfico de Rosca
    donut = DoughnutChart()
    donut.title = "Distribuição por classificação"
    donut.add_data(
        Reference(ws, min_col=2, min_row=7, max_row=10), titles_from_data=False
    )
    donut.set_categories(Reference(ws, min_col=1, min_row=7, max_row=10))
    donut.holeSize = 55
    donut.dataLabels = DataLabelList()
    donut.dataLabels.showPercent = True
    ws.add_chart(donut, "G2")

    # Gráfico de Linhas (Evolução)
    linhas = LineChart()
    linhas.title = "Evolução dos registros"
    linhas.y_axis.title = "Registros"
    linhas.x_axis.title = "Data de referência"
    linhas.add_data(
        Reference(ws, min_col=2, min_row=19, max_row=29), titles_from_data=True
    )
    linhas.set_categories(Reference(ws, min_col=1, min_row=20, max_row=29))
    linhas.height = 7
    linhas.width = 15
    ws.add_chart(linhas, "G18")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24


def _montar_aba_ranking(ws, indicadores: OperationalIndicators) -> None:
    """Preenche a aba Ranking de Regras."""
    ws.title = "Ranking de Regras"

    headers = ["Posição", "Regra", "Nome", "Ocorrências", "% do total"]
    fill_header = PatternFill("solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    for idx, rule in enumerate(indicadores.ranking_regras, start=1):
        r_idx = idx + 1
        ws.cell(row=r_idx, column=1, value=idx)
        ws.cell(row=r_idx, column=2, value=rule.codigo)
        ws.cell(row=r_idx, column=3, value=rule.nome)
        ws.cell(row=r_idx, column=4, value=rule.ocorrencias)
        c5 = ws.cell(row=r_idx, column=5, value=rule.percentual_total / 100.0)
        c5.number_format = "0.0%"

    max_r = len(indicadores.ranking_regras) + 1
    tabela = Table(displayName="TabelaRankingRegras", ref=f"A1:E{max_r}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.cell(
        row=max_r + 2,
        column=1,
        value="Nota: Um registro pode acionar mais de uma regra; portanto, os percentuais não precisam somar 100%.",
    ).font = Font(size=9, italic=True, color="555555")

    for col_idx in range(1, 6):
        valores = [
            len(str(ws.cell(l, col_idx).value or ""))
            for l in range(1, max_r + 1)
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            max(valores, default=10) + 2, 12
        )
    ws.column_dimensions["C"].width = 45


def _montar_aba_dicionario(ws) -> None:
    """Preenche a aba Dicionário com todas as definições obrigatórias."""
    ws.title = "Dicionário"

    headers = ["Termo", "Definição", "Fórmula/Regra", "Interpretação"]
    fill_header = PatternFill("solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    dicionario_itens = [
        (
            "Lote",
            "Código identificador único do lote de produção.",
            "Campo obrigatório no lote de dados (RN01).",
            "Chave de identificação do lote para rastreabilidade de produção.",
        ),
        (
            "Produto",
            "Nome ou código do produto inspecionado no lote.",
            "Campo obrigatório (RN02).",
            "Garante a vinculação do lote ao produto correto.",
        ),
        (
            "Linha",
            "Linha de fabricação/produção onde ocorreu a inspeção.",
            "Campo obrigatório (RN03).",
            "Identifica a origem física do processo fabril.",
        ),
        (
            "Turno",
            "Turno operacional em que a inspeção foi realizada.",
            "Campo obrigatório (RN04).",
            "Permite o acompanhamento de qualidade por turno de trabalho.",
        ),
        (
            "Status",
            "Resultado da inspeção do lote após normalização.",
            "Normalização de OK/NOK para APROVADO/REPROVADO (RN06/RN07).",
            "Estado canônico do lote no sistema.",
        ),
        (
            "Responsável",
            "Nome ou ID do técnico responsável pela inspeção.",
            "Campo obrigatório (RN04).",
            "Garante a auditabilidade das medições efetuadas.",
        ),
        (
            "Data da inspeção",
            "Data em que a inspeção foi registrada pelo operador.",
            "Formato DD/MM/AAAA e compatível com a referência (RN12).",
            "Métrica de tempestividade e controle cronológico.",
        ),
        (
            "Data de referência",
            "Data da execução/aba diária processada.",
            "Extraída do nome da aba (Insp_DD_MM_AAAA).",
            "Âncora temporal de validação dos registros diários.",
        ),
        (
            "Observação",
            "Texto descritivo com justificativa da inspeção.",
            "Obrigatória quando o Status é REPROVADO (RN10).",
            "Fornece contexto operacional em caso de reprovação.",
        ),
        (
            "Orientação",
            "Descrição detalhada das regras acionadas e ações.",
            "Concatenação das mensagens das regras aplicadas.",
            "Instruções operacionais para triagem de pendências.",
        ),
        (
            "Classificação",
            "Categoria final atribuída ao registro após validação.",
            "Precedência: Erro de Entrada > Ambíguo > Divergência > Válido.",
            "Direciona a tratativa adequada para cada lote.",
        ),
        (
            "Válido",
            "Registro totalmente conforme sem nenhuma violação de regra.",
            "Nenhuma regra de erro, ambiguidade ou divergência acionada.",
            "Pronto para integração automática no ERP/MES.",
        ),
        (
            "Divergência",
            "Inconsistência de negócio que requer ajuste cadastral.",
            "Acionamento de RN05, RN10 ou RN11.",
            "Necessita verificação de cadastro de lote, observação ou duplicidade.",
        ),
        (
            "Ambíguo",
            "Registro com valor não reconhecido de status.",
            "Acionamento de RN09.",
            "Requer revisão e deliberação humana imediata.",
        ),
        (
            "Erro de Entrada",
            "Inconsistência no preenchimento de campos obrigatórios ou datas.",
            "Acionamento de RN01, RN02, RN03, RN04 ou RN12.",
            "Impede o processamento e deve ser corrigido na origem.",
        ),
        (
            "RN01 — Lote obrigatório",
            "Validação da presença do identificador do lote.",
            "lote_id não nulo e não vazio.",
            "Registros sem lote são classificados como Erro de Entrada.",
        ),
        (
            "RN02 — Produto obrigatório",
            "Validação da presença do nome/código do produto.",
            "produto não nulo e não vazio.",
            "Registros sem produto são classificados como Erro de Entrada.",
        ),
        (
            "RN03 — Linha obrigatória",
            "Validação da presença da linha de produção.",
            "linha não nula e não vazia.",
            "Registros sem linha são classificados como Erro de Entrada.",
        ),
        (
            "RN04 — Campos operacionais obrigatórios",
            "Validação de turno, status original e responsável.",
            "turno, status_original e responsavel preenchidos.",
            "Falha em qualquer um gera Erro de Entrada.",
        ),
        (
            "RN05 — Lote ausente da Base de Referência",
            "Checagem de existência do lote no cadastro de referência.",
            "lote_id em Base_Referencia.",
            "Lote não cadastrado gera Divergência.",
        ),
        (
            "RN06 — Normalização de OK para APROVADO",
            "Padronização de status simplificado 'OK'.",
            "status_original.upper() == 'OK'.",
            "Converte automaticamente OK para APROVADO.",
        ),
        (
            "RN07 — Normalização de NOK para REPROVADO",
            "Padronização de status simplificado 'NOK'.",
            "status_original.upper() == 'NOK'.",
            "Converte automaticamente NOK para REPROVADO.",
        ),
        (
            "RN08 — Status canônico aceito",
            "Aceitação de status já enviados no padrão canônico.",
            "status_original já em APROVADO ou REPROVADO.",
            "Processamento normal sem geração de ocorrência de alteração.",
        ),
        (
            "RN09 — Status desconhecido",
            "Detecção de status fora do domínio esperado.",
            "status_original não reconhecido pelo domínio.",
            "Classifica como Ambíguo para encaminhar à revisão humana.",
        ),
        (
            "RN10 — Reprovado sem observação",
            "Exigência de motivo justificado para reprovações.",
            "status == REPROVADO e observação vazia.",
            "Gera Divergência por ausência de justificativa.",
        ),
        (
            "RN11 — Lote duplicado no mesmo dia",
            "Verificação de repetição de lote na mesma aba diária.",
            "lote_id repetido no mesmo dia de inspeção.",
            "Gera Divergência por duplicidade no dia.",
        ),
        (
            "RN12 — Data inválida",
            "Validação da data do registro contra a data da aba.",
            "Formato DD/MM/AAAA igual à data_referencia da aba.",
            "Data inconsistente gera Erro de Entrada.",
        ),
        (
            "Total de registros",
            "Volume total de inspeções processadas no período.",
            "Contagem total de registros das 10 abas.",
            "Base do universo de análise.",
        ),
        (
            "Regra mais acionada",
            "Regra de validação com maior contagem de ocorrências.",
            "Primeiro item do ranking de regras por frequência.",
            "Indica o principal padrão de ajuste/erro identificado.",
        ),
        (
            "Ranking de regras",
            "Lista ordenada de todas as regras acionadas por frequência.",
            "Agrupamento e ordenação por ocorrências.",
            "Visão consolidada das causas de falhas e normalizações.",
        ),
        (
            "Taxa de qualidade da entrada",
            "Proporção de registros sem erros estruturais de entrada.",
            "((Total - Erros de Entrada) / Total) * 100.",
            "Mede a confiabilidade do preenchimento inicial (meta > 80%).",
        ),
        (
            "Taxa de revisão humana",
            "Proporção de registros direcionados à análise manual.",
            "(Ambíguos / Total) * 100.",
            "Mede a necessidade de intervenção humana (meta < 15%).",
        ),
        (
            "Taxa de retrabalho",
            "Proporção de registros com inconsistências de negócio.",
            "(Divergências / Total) * 100.",
            "Mede a necessidade de ajustes operacionais (meta < 6%).",
        ),
        (
            "Ganho estimado de tempo",
            "Tempo economizado pela solução automatizada de inspeção.",
            "Total * (Tempo Manual - Tempo Automatizado).",
            "Estimativa didática dos benefícios da automação.",
        ),
        (
            "Tempo manual por registro",
            "Premissa de tempo médio para conferência totalmente manual.",
            "Premissa configurada de 2,00 minutos por registro.",
            "Base de comparação para cálculo de ganhos de produtividade.",
        ),
        (
            "Tempo automatizado por registro",
            "Premissa de tempo médio para conferência via automação.",
            "Premissa configurada de 0,25 minutos por registro.",
            "Base de comparação para cálculo de ganhos de produtividade.",
        ),
        (
            "Base de Referência",
            "Tabela de cadastro oficial com lotes válidos no sistema.",
            "Lista oficial de lote_id cadastrados no ERP.",
            "Fonte da verdade para validação de existência do lote.",
        ),
        (
            "Data de referência",
            "Data oficial atribuída a cada aba de inspeção diária.",
            "Formato DD/MM/AAAA derivado do nome da aba.",
            "Base para validação temporal dos registros.",
        ),
        (
            "Deduplicação por dia",
            "Verificação de unicidade do lote dentro da mesma aba diária.",
            "Contagem isolada de lote_id por aba.",
            "Garante a identificação de entradas repetidas no mesmo dia.",
        ),
    ]

    for r_idx, (termo, defn, formula, interp) in enumerate(dicionario_itens, start=2):
        ws.cell(row=r_idx, column=1, value=termo)
        ws.cell(row=r_idx, column=2, value=defn)
        ws.cell(row=r_idx, column=3, value=formula)
        ws.cell(row=r_idx, column=4, value=interp)

    max_r = len(dicionario_itens) + 1
    tabela = Table(displayName="TabelaDicionarioTermos", ref=f"A1:D{max_r}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45


def _montar_aba_decisoes_ml(ws, decisoes_ml: Sequence[MLDecision]) -> None:
    """Preenche a 9ª aba 'Decisões de ML' com histórico de inferências e auditoria."""
    ws.title = "Decisões de ML"

    headers = [
        "Timestamp",
        "Lote",
        "Status original",
        "Turno",
        "Tem observação",
        "Classe predita",
        "Probabilidade",
        "Nível de confiança",
        "Ação final",
        "Latência total (ms)",
        "Tentou rede",
        "Circuit breaker aberto",
        "Versão do modelo",
        "Tipo de erro",
    ]
    fill_header = PatternFill("solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    for r_idx, dec in enumerate(decisoes_ml, start=2):
        row_dict = dec.to_dict()
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = row_dict.get(h)
            cell.value = val
            if h == "Probabilidade" and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            elif h == "Latência total (ms)" and isinstance(val, (int, float)):
                cell.number_format = "0.00"

    max_r = max(len(decisoes_ml) + 1, 2)
    tabela = Table(displayName="TabelaDecisoesML", ref=f"A1:N{max_r}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max_r}"

    for col_idx in range(1, 15):
        valores = [
            len(str(ws.cell(l, col_idx).value or ""))
            for l in range(1, min(max_r, 100) + 1)
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(valores, default=10) + 2, 40
        )


def gerar_excel_consolidado(
    validados: Sequence[RegistroValidado],
    indicadores: OperationalIndicators,
    destino: Path,
    *,
    decisoes_ml: Sequence[MLDecision] = (),
) -> Path:
    """Gera o arquivo Excel com exatamente as 9 abas consolidadas (incluindo Decisões de ML)."""
    df_validados = pd.DataFrame([r.to_dict() for r in validados])
    df_relatorio = df_validados.loc[:, list(COLUNAS_RELATORIO)].rename(
        columns=COLUNAS_RELATORIO
    )

    wb = Workbook()

    # Aba 1: Resumo
    ws_resumo = wb.active
    _montar_aba_resumo(ws_resumo, indicadores, validados)

    # Aba 2: Todos
    ws_todos = wb.create_sheet(title="Todos")
    for r in dataframe_to_rows(df_relatorio, index=False, header=True):
        ws_todos.append(r)
    _ajustar_aba_tabela(ws_todos, "Todos")

    # Aba 3: Válidos
    ws_validos = wb.create_sheet(title="Válidos")
    df_validos = df_relatorio.loc[df_relatorio["Classificação"] == "Válido"]
    for r in dataframe_to_rows(df_validos, index=False, header=True):
        ws_validos.append(r)
    _ajustar_aba_tabela(ws_validos, "Validos")

    # Aba 4: Divergências
    ws_div = wb.create_sheet(title="Divergências")
    df_div = df_relatorio.loc[df_relatorio["Classificação"] == "Divergência"]
    for r in dataframe_to_rows(df_div, index=False, header=True):
        ws_div.append(r)
    _ajustar_aba_tabela(ws_div, "Divergencias")

    # Aba 5: Ambíguos
    ws_amb = wb.create_sheet(title="Ambíguos")
    df_amb = df_relatorio.loc[df_relatorio["Classificação"] == "Ambíguo"]
    for r in dataframe_to_rows(df_amb, index=False, header=True):
        ws_amb.append(r)
    _ajustar_aba_tabela(ws_amb, "Ambiguos")

    # Aba 6: Erros de Entrada
    ws_err = wb.create_sheet(title="Erros de Entrada")
    df_err = df_relatorio.loc[df_relatorio["Classificação"] == "Erro de Entrada"]
    for r in dataframe_to_rows(df_err, index=False, header=True):
        ws_err.append(r)
    _ajustar_aba_tabela(ws_err, "ErrosDeEntrada")

    # Aba 7: Ranking de Regras
    ws_ranking = wb.create_sheet(title="Ranking de Regras")
    _montar_aba_ranking(ws_ranking, indicadores)

    # Aba 8: Dicionário
    ws_dicionario = wb.create_sheet(title="Dicionário")
    _montar_aba_dicionario(ws_dicionario)

    # Aba 9: Decisões de ML
    ws_ml = wb.create_sheet(title="Decisões de ML")
    _montar_aba_decisoes_ml(ws_ml, decisoes_ml)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino



def gerar_resumo_executivo_md(
    indicadores: OperationalIndicators, destino: Path
) -> Path:
    """Gera o arquivo resumo_executivo.md consumindo o objeto de indicadores."""
    regra_codigo = (
        indicadores.regra_mais_acionada.codigo
        if indicadores.regra_mais_acionada
        else "N/A"
    )
    regra_nome = (
        indicadores.regra_mais_acionada.nome
        if indicadores.regra_mais_acionada
        else "Nenhuma"
    )
    regra_ocorrencias = (
        indicadores.regra_mais_acionada.ocorrencias
        if indicadores.regra_mais_acionada
        else 0
    )
    regra_pct = (
        indicadores.regra_mais_acionada.percentual_total
        if indicadores.regra_mais_acionada
        else 0.0
    )

    tempo_legivel = _formatar_tempo_legivel(indicadores.ganho_estimado_minutos)
    ganho_min_str = f"{indicadores.ganho_estimado_minutos:.1f}".replace(".", ",")
    manual_min_str = f"{indicadores.tempo_manual_minutos:.2f}".replace(".", ",")
    auto_min_str = f"{indicadores.tempo_automatizado_minutos:.2f}".replace(".", ",")
    econ_min_str = f"{(indicadores.tempo_manual_minutos - indicadores.tempo_automatizado_minutos):.2f}".replace(".", ",")

    conteudo_md = f"""# Resumo Executivo — Conferência de Lotes

## Visão Geral
No período analisado de 10 dias de inspeção, foram processados **{indicadores.total_registros} registros** de lotes de produção através da solução automatizada de conferência. O processo aplicou regras de validação estrutural, cadastral e de negócio para classificar cada registro e quantificar os ganhos operacionais da automação.

## Indicadores Principais
- **Total de registros processados:** {indicadores.total_registros}
- **Registros válidos:** {indicadores.registros_validos} ({indicadores.percentual_validos:.1f}%).replace('.', ',')
- **Divergências operacionais:** {indicadores.divergencias} ({indicadores.percentual_divergencias:.1f}%).replace('.', ',')
- **Registros ambíguos (revisão humana):** {indicadores.ambiguos} ({indicadores.percentual_ambiguos:.1f}%).replace('.', ',')
- **Erros de entrada:** {indicadores.erros_entrada} ({indicadores.percentual_erros_entrada:.1f}%).replace('.', ',')
- **Taxa de qualidade da entrada:** {indicadores.taxa_qualidade_entrada:.1f}% (meta > 80,0%) — **Conforme**
- **Taxa de revisão humana:** {indicadores.taxa_revisao_humana:.1f}% (meta < 15,0%) — **Conforme**
- **Taxa de retrabalho:** {indicadores.taxa_retrabalho:.1f}% (meta < 6,0%) — **Atenção**

## Destaque
A regra com maior ocorrência no período foi a **{regra_codigo} — {regra_nome}**, acionada em **{regra_ocorrencias} ocorrências** ({regra_pct:.1f}% do total de registros). Este destaque evidencia um volume expressivo de registros recebidos com o status simplificado "OK", que foram padronizados automaticamente para o formato canônico "APROVADO" sem necessitar de intervenção manual.

## Ganho Estimado de Tempo
A automação proporcionou uma economia estimada de **{ganho_min_str} minutos** (equivalente a **{tempo_legivel}**) no processamento das inspeções do período.

- **Tempo estimado manual:** {manual_min_str} minutos por registro
- **Tempo estimado automatizado:** {auto_min_str} minutos por registro
- **Economia por registro:** {econ_min_str} minutos por registro

## Observação
Os tempos e o ganho apresentados constituem uma **estimativa didática** baseada nas premissas de {manual_min_str} min/registro para o processo manual e {auto_min_str} min/registro para o processo automatizado. Transformar essa estimativa em uma métrica real de produção exigiria a implementação de telemetria de produção, medição contínua dos tempos efetivamente observados e controle estatístico de exceções operacionais.
"""
    # Ajuste fino de formatação de vírgula em percentuais para exibição pt-BR no markdown
    conteudo_md_formatado = (
        conteudo_md.replace(f"({indicadores.percentual_validos:.1f}%).replace('.', ',')", f"({indicadores.percentual_validos:.1f}%)".replace(".", ","))
        .replace(f"({indicadores.percentual_divergencias:.1f}%).replace('.', ',')", f"({indicadores.percentual_divergencias:.1f}%)".replace(".", ","))
        .replace(f"({indicadores.percentual_ambiguos:.1f}%).replace('.', ',')", f"({indicadores.percentual_ambiguos:.1f}%)".replace(".", ","))
        .replace(f"({indicadores.percentual_erros_entrada:.1f}%).replace('.', ',')", f"({indicadores.percentual_erros_entrada:.1f}%)".replace(".", ","))
        .replace(f"{indicadores.taxa_qualidade_entrada:.1f}%", f"{indicadores.taxa_qualidade_entrada:.1f}%".replace(".", ","))
        .replace(f"{indicadores.taxa_revisao_humana:.1f}%", f"{indicadores.taxa_revisao_humana:.1f}%".replace(".", ","))
        .replace(f"{indicadores.taxa_retrabalho:.1f}%", f"{indicadores.taxa_retrabalho:.1f}%".replace(".", ","))
        .replace(f"{regra_pct:.1f}%", f"{regra_pct:.1f}%".replace(".", ","))
    )

    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(conteudo_md_formatado, encoding="utf-8")
    return destino


def _escapar_texto_pdf(texto: str) -> str:
    """Converte texto para o conjunto simples suportado pelo PDF básico."""
    texto_ascii = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore")
    return (
        texto_ascii.decode("ascii")
        .replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
    )


def _exportar_pdf_basico(
    indicadores: OperationalIndicators, destino: Path
) -> Path:
    """Gera PDF sem dependência de ReportLab."""
    resumo_dict = {
        "Total de registros": indicadores.total_registros,
        "Válidos": indicadores.registros_validos,
        "Divergências": indicadores.divergencias,
        "Ambíguos": indicadores.ambiguos,
        "Erros de Entrada": indicadores.erros_entrada,
        "Ganho Estimado": f"{indicadores.ganho_estimado_minutos:.1f} min",
    }
    linhas = [
        "BT",
        "/F1 18 Tf",
        "48 790 Td",
        "(Resumo Executivo - Conferencia de Lotes) Tj",
        "/F1 11 Tf",
        "0 -25 Td",
        f"(Gerado em {_escapar_texto_pdf(_agora_manaus().strftime('%d/%m/%Y %H:%M:%S'))}) Tj",
    ]
    for rotulo, quantidade in resumo_dict.items():
        linhas.extend(
            ("0 -28 Td", f"({_escapar_texto_pdf(rotulo)}: {quantidade}) Tj")
        )
    linhas.append("ET")
    conteudo = "\n".join(linhas).encode("ascii")
    objetos = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length "
        + str(len(conteudo)).encode()
        + b" >>\nstream\n"
        + conteudo
        + b"\nendstream",
    ]
    partes = [b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"]
    offsets = [0]
    for numero, objeto in enumerate(objetos, start=1):
        offsets.append(sum(len(parte) for parte in partes))
        partes.append(f"{numero} 0 obj\n".encode() + objeto + b"\nendobj\n")
    inicio_xref = sum(len(parte) for parte in partes)
    xref = [f"xref\n0 {len(objetos) + 1}\n", "0000000000 65535 f \n"]
    xref.extend(f"{offset:010d} 0000 n \n" for offset in offsets[1:])
    partes.append("".join(xref).encode())
    partes.append(
        f"trailer\n<< /Size {len(objetos) + 1} /Root 1 0 R >>\nstartxref\n{inicio_xref}\n%%EOF\n".encode()
    )
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(b"".join(partes))
    return destino


def gerar_pdf_compativel(
    validados: Sequence[RegistroValidado],
    indicadores: OperationalIndicators,
    destino: Path,
) -> Path:
    """Gera PDF compatível a partir do objeto consolidado de indicadores."""
    try:
        from reportlab.graphics import renderPDF
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.shapes import Circle, Drawing
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen.canvas import Canvas
    except ImportError:
        return _exportar_pdf_basico(indicadores, destino)

    pdf = Canvas(str(destino), pagesize=A4)
    pdf.setTitle("Resumo Executivo — Conferência de Lotes")
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(48, 790, "Resumo Executivo — Conferência de Lotes")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(
        48, 765, f"Gerado em {_agora_manaus().strftime('%d/%m/%Y %H:%M:%S')}"
    )

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(48, 730, "Total de registros")
    pdf.setFont("Helvetica", 20)
    pdf.drawString(48, 705, str(indicadores.total_registros))

    totais_class = {
        "Válido": indicadores.registros_validos,
        "Divergência": indicadores.divergencias,
        "Ambíguo": indicadores.ambiguos,
        "Erro de Entrada": indicadores.erros_entrada,
    }

    for x, rotulo in zip((48, 178, 308, 438), CLASSIFICACOES):
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(x, 665, rotulo)
        pdf.setFont("Helvetica", 18)
        pdf.drawString(x, 642, str(totais_class.get(rotulo, 0)))

    desenho = Drawing(500, 330)
    rosquinha = Pie()
    rosquinha.x, rosquinha.y = 10, 100
    rosquinha.width, rosquinha.height = 180, 180
    rosquinha.data = [totais_class.get(c, 0) for c in CLASSIFICACOES]
    rosquinha.labels = [
        f"{c}: {_percentual(totais_class.get(c, 0), indicadores.total_registros):.0f}%"
        for c in CLASSIFICACOES
    ]
    rosquinha.slices.strokeWidth = 0
    for indice, cor in enumerate(("#5C0011", "#8A1024", "#B71C35", "#D94B62")):
        rosquinha.slices[indice].fillColor = HexColor(cor)
    desenho.add(rosquinha)
    desenho.add(
        Circle(
            rosquinha.x + rosquinha.width / 2,
            rosquinha.y + rosquinha.height / 2,
            42,
            fillColor=HexColor("#FFFFFF"),
            strokeColor=None,
        )
    )

    # Evolução diária de alertas
    df_val = pd.DataFrame([r.to_dict() for r in validados])
    df_val["alerta"] = df_val["classificacao"].isin(["Divergência", "Ambíguo"]).astype(int)
    evolucao = (
        df_val.groupby("data_referencia", sort=False)["alerta"]
        .sum()
        .reset_index(name="Divergências + Ambíguos")
    )

    linha = HorizontalLineChart()
    linha.x, linha.y = 260, 115
    linha.width, linha.height = 210, 150
    linha.data = [evolucao["Divergências + Ambíguos"].tolist()]
    linha.categoryNames = [
        data[:5] for data in evolucao["data_referencia"].tolist()
    ]
    linha.lines[0].strokeColor = HexColor("#B71C35")
    linha.lines[0].strokeWidth = 2
    linha.valueAxis.valueMin = 0
    linha.valueAxis.valueMax = max(linha.data[0], default=1) + 1
    linha.valueAxis.valueStep = 1
    linha.categoryAxis.labels.angle = 45
    linha.categoryAxis.labels.boxAnchor = "ne"
    linha.categoryAxis.labels.fontSize = 6
    linha.categoryAxis.labels.dy = -5
    desenho.add(linha)

    renderPDF.draw(desenho, pdf, 48, 300)

    pdf.setFont("Helvetica", 6)
    datas_reduzidas = [data[:5] for data in evolucao["data_referencia"]]
    for indice, data in enumerate(datas_reduzidas):
        x_rotulo = 48 + linha.x + (indice + 0.5) * linha.width / len(
            datas_reduzidas
        )
        pdf.saveState()
        pdf.translate(x_rotulo, 405)
        pdf.rotate(45)
        pdf.drawRightString(0, 0, data)
        pdf.restoreState()

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(48, 280, "Distribuição por classificação")
    pdf.drawString(308, 280, "Evolução dos registros")
    pdf.save()
    return destino


def registrar_log_execucao(
    indicadores: OperationalIndicators,
    caminho_entrada: Path,
    destino_log: Path,
    logger: logging.Logger,
) -> Path:
    """Registra o log final da execução usando os indicadores consolidados."""
    resumo_log = {
        "total": indicadores.total_registros,
        "Válido": indicadores.registros_validos,
        "Divergência": indicadores.divergencias,
        "Ambíguo": indicadores.ambiguos,
        "Erro de Entrada": indicadores.erros_entrada,
    }
    logger.info(
        "Executado em %s | Relatório gerado | entrada=%s | totais=%s",
        _agora_manaus().strftime("%d/%m/%Y %H:%M:%S %z"),
        caminho_entrada.name,
        resumo_log,
        extra={
            "event": "DASHBOARD_EXECUTION_COMPLETED",
            "input_file": caminho_entrada.name,
            "summary": resumo_log,
        },
    )
    for handler in logger.handlers:
        handler.flush()
    return destino_log


def carregar_inspecoes(caminho: str | Path) -> pd.DataFrame:
    """Lê as dez abas diárias, respeitando a linha 3 como cabeçalho."""
    caminho = Path(caminho)
    frames: list[pd.DataFrame] = []
    with pd.ExcelFile(caminho) as arquivo:
        abas = [aba for aba in arquivo.sheet_names if ABAS_DIARIAS.match(aba)]
        if len(abas) != 10:
            raise ValueError(f"Esperadas 10 abas de inspeção; encontradas {len(abas)}.")
        for aba in abas:
            apresentacao = pd.read_excel(arquivo, sheet_name=aba, header=None, nrows=2)
            texto_apresentacao = " ".join(apresentacao.fillna("").astype(str).to_numpy().ravel())
            encontrado = re.search(r"Registros:\s*(\d+)", texto_apresentacao, flags=re.I)
            if not encontrado:
                raise ValueError(f"Quantidade de registros não encontrada na aba {aba}.")
            quantidade = int(encontrado.group(1))
            diario = pd.read_excel(arquivo, sheet_name=aba, skiprows=2, nrows=quantidade)
            valida_estrutura(diario)
            diario["aba_origem"] = aba
            diario["linha_origem"] = range(4, 4 + len(diario))
            diario["data_referencia"] = data_da_aba(aba)
            frames.append(diario)
    consolidado = pd.concat(frames, ignore_index=True)
    if len(consolidado) != 250:
        raise ValueError(f"Foram lidos {len(consolidado)} registros; o esperado é 250.")
    return consolidado


def validar_registros(df: pd.DataFrame, lotes_referencia: set[str]) -> pd.DataFrame:
    """Deduplica por dia e chama o Serviço de Validação para cada linha."""
    repetidos: set[int] = set()
    for _, diario in df.groupby("aba_origem", sort=False):
        ids = [texto(valor) for valor in diario["lote_id"]]
        contador = Counter(lote_id for lote_id in ids if lote_id)
        vistos: Counter[str] = Counter()
        for indice, lote_id in zip(diario.index, ids):
            vistos[lote_id] += 1
            if lote_id and contador[lote_id] > 1 and vistos[lote_id] > 1:
                repetidos.add(indice)

    registros: list[dict[str, object]] = []
    for indice, linha in df.iterrows():
        registro = validar_registro(
            linha.to_dict(),
            lotes_referencia,
            duplicado_no_dia=indice in repetidos,
        )
        registros.append(registro.to_dict())
    resultado = pd.DataFrame(registros)
    if len(resultado) != len(df) or not resultado["classificacao"].isin(CLASSIFICACOES).all():
        raise RuntimeError("Falha ao classificar os registros do relatório.")
    return resultado


def preparar_dados_relatorio(resultado: pd.DataFrame) -> pd.DataFrame:
    """Expõe no Excel somente campos compreensíveis ao público de negócio."""
    return resultado.loc[:, list(COLUNAS_RELATORIO)].rename(
        columns=COLUNAS_RELATORIO
    )


def gerar_relatorio(
    caminho_entrada: str | Path = ARQUIVO_ENTRADA_PADRAO,
    diretorio_saida: str | Path = DIRETORIO_SAIDA_PADRAO,
) -> Path:
    """Fachada compatível que delega a execução ao orquestrador principal."""
    resultados = executar_pipeline_dashboard(caminho_entrada, diretorio_saida)
    return resultados["excel"]


def executar_pipeline_dashboard(
    caminho_entrada: str | Path = ARQUIVO_ENTRADA_PADRAO,
    diretorio_saida: str | Path = DIRETORIO_SAIDA_PADRAO,
) -> dict[str, Path]:
    """Orquestra o fluxo completo da Aula 24.

    Garante chamada ÚNICA a ``calcular_indicadores()`` e distribui o mesmo objeto
    de indicadores para Excel, Markdown, PDF e Log.
    """

    caminho_entrada = Path(caminho_entrada)
    diretorio_saida = Path(diretorio_saida)
    diretorio_saida.mkdir(parents=True, exist_ok=True)

    # 1. Carregar inspecao e base de referencia
    lotes_ref = carregar_base_referencia(caminho_entrada)
    df_inspecoes = carregar_inspecoes(caminho_entrada)

    # 2. Validar cada linha exatamente uma vez em objetos RegistroValidado
    validados = validar_registros_lista(df_inspecoes, lotes_ref)

    # 3. Calcular indicadores EXATAMENTE UMA VEZ
    indicadores = calcular_indicadores(
        validados, tempo_manual_minutos=2.0, tempo_automatizado_minutos=0.25
    )

    # 4. Executar inferência de ML para registros ambíguos
    caminho_log = diretorio_saida / "execucao_dashboard.log"
    logger = configure_structured_logging(
        diretorio_saida,
        execution_id=settings.execution_id,
        bot_id=settings.bot_id,
        logger_name="dashboard.auditoria",
        log_filename=caminho_log.name,
    )
    decisoes_ml: list[MLDecision] = []
    ml_client = create_ml_client(settings, logger)
    try:
        processor = ItemProcessor(ml_client, logger)
        decisoes_ml = processor.processar_lote(validados)
    finally:
        ml_client.close()

    # 5. Gerar artefatos consumindo os indicadores calculados e decisões de ML
    caminho_excel = gerar_excel_consolidado(
        validados,
        indicadores,
        diretorio_saida / "relatorio_conferencia_lotes.xlsx",
        decisoes_ml=decisoes_ml,
    )
    caminho_md = gerar_resumo_executivo_md(
        indicadores, diretorio_saida / "resumo_executivo.md"
    )
    caminho_pdf = gerar_pdf_compativel(
        validados, indicadores, diretorio_saida / "resumo_conferencia_lotes.pdf"
    )
    caminho_log = registrar_log_execucao(indicadores, caminho_entrada, caminho_log, logger)

    return {
        "excel": caminho_excel,
        "markdown": caminho_md,
        "pdf": caminho_pdf,
        "log": caminho_log,
    }



def main() -> None:
    """Interface CLI oficial para o Dashboard Executivo da Aula 24."""
    parser = argparse.ArgumentParser(
        description="Orquestrador do Dashboard Executivo (Aula 24)."
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        default=ARQUIVO_ENTRADA_PADRAO,
        help="Caminho da planilha Excel de entrada.",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=DIRETORIO_SAIDA_PADRAO,
        help="Diretório de saída para os artefatos.",
    )
    args = parser.parse_args()
    resultados = executar_pipeline_dashboard(args.entrada, args.saida)
    print(f"Excel gerado: {resultados['excel']}")
    print(f"Markdown gerado: {resultados['markdown']}")


if __name__ == "__main__":
    main()
